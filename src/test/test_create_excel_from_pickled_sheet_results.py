import argparse
import pickle
import shutil
import tempfile
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from typing import Iterator

import pytest
from openpyxl import load_workbook

import src.create_excel_from_pickled_sheet_results as pickle_to_excel
from src.create_excel_from_pickled_sheet_results import (
    apply_duplicate_model_labels,
    build_excel_filename,
    discover_sheet_results_pickle_paths,
    filter_sheet_results_by_target_modes,
    load_pickled_sheet_results,
    load_sheet_results_from_dir,
    save_excel_summaries_from_sheet_results,
)
from src.create_exhaustive_study_sheet import simulate_cross_val_summary
from src.schemas.dataclasses import SheetResult
from src.settings import CLI_CHOICES, PATHS


@pytest.fixture
def local_tmp_path() -> Iterator[Path]:
    base_dir = Path(__file__).resolve().parents[1] / "pytest_local_tmp"
    base_dir.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(tempfile.mkdtemp(prefix="pickle_excel_", dir=base_dir))
    try:
        yield temp_dir
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _build_args(
    *,
    dataset_v: str = "uk",
    model_type: str = "test_cnn",
    recording_category: str = "poem",
    selected_classes: str = "copd,control",
    norm_mode: str = "norm_off",
    add_demo_data: bool = False,
) -> argparse.Namespace:
    return argparse.Namespace(
        dataset_v=dataset_v,
        model_type=model_type,
        recording_category=recording_category,
        selected_classes=selected_classes,
        norm_mode=norm_mode,
        add_demo_data=add_demo_data,
        test_run=False,
    )


def _build_sheet_result(
    *,
    dataset_v: str = "uk",
    model_type: str = "test_cnn",
    recording_category: str = "poem",
    selected_classes: str = "copd,control",
    norm_mode: str = "norm_off",
    add_demo_data: bool = False,
) -> SheetResult:
    args = _build_args(
        dataset_v=dataset_v,
        model_type=model_type,
        recording_category=recording_category,
        selected_classes=selected_classes,
        norm_mode=norm_mode,
        add_demo_data=add_demo_data,
    )
    return simulate_cross_val_summary(args, nan_probability=0.0)


def _write_pickle(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as file_handle:
        pickle.dump(payload, file_handle)


EXPORT_TIMESTAMP = "2026-03-19_14-05-33"


@pytest.mark.unit
class TestCreateExcelFromPickledSheetResults:
    def test_discover_sheet_results_pickle_paths_returns_sorted_matching_files_only(
        self,
        local_tmp_path: Path,
    ) -> None:
        later_match = local_tmp_path / f"b{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}"
        earlier_match = local_tmp_path / f"a{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}"
        ignored = local_tmp_path / "ignore.pkl"

        later_match.write_bytes(b"pickle")
        earlier_match.write_bytes(b"pickle")
        ignored.write_bytes(b"pickle")

        discovered = discover_sheet_results_pickle_paths(local_tmp_path)

        assert discovered == [earlier_match, later_match]

    def test_discover_sheet_results_pickle_paths_accepts_variant_aware_names(
        self,
        local_tmp_path: Path,
    ) -> None:
        later_match = (
            local_tmp_path
            / f"study_b_norm_on_uk_no_tuning_w_demo_data_2fold{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}"
        )
        earlier_match = (
            local_tmp_path
            / f"study_a_norm_off_uk_no_tuning_no_demo_data_2fold{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}"
        )
        ignored = local_tmp_path / "ignore_sheet_results.txt"

        later_match.write_bytes(b"pickle")
        earlier_match.write_bytes(b"pickle")
        ignored.write_bytes(b"pickle")

        discovered = discover_sheet_results_pickle_paths(local_tmp_path)

        assert discovered == [earlier_match, later_match]

    def test_discover_sheet_results_pickle_paths_raises_when_no_matching_files(
        self,
        local_tmp_path: Path,
    ) -> None:
        (local_tmp_path / "ignore.pkl").write_bytes(b"pickle")

        with pytest.raises(ValueError, match="sheet_results"):
            discover_sheet_results_pickle_paths(local_tmp_path)

    def test_load_pickled_sheet_results_rejects_non_list_payload(
        self,
        local_tmp_path: Path,
    ) -> None:
        pickle_path = local_tmp_path / f"invalid{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}"
        _write_pickle(pickle_path, {"not": "a list"})

        with pytest.raises(ValueError, match="list\\[SheetResult\\]"):
            load_pickled_sheet_results(pickle_path)

    def test_load_sheet_results_from_dir_combines_all_matching_pickles(
        self,
        local_tmp_path: Path,
    ) -> None:
        first_result = _build_sheet_result(recording_category="poem", add_demo_data=False)
        second_result = _build_sheet_result(recording_category="a", add_demo_data=False)
        _write_pickle(local_tmp_path / f"a{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}", [first_result])
        _write_pickle(local_tmp_path / f"b{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}", [second_result])
        _write_pickle(local_tmp_path / "ignored.pkl", [first_result])

        combined = load_sheet_results_from_dir(local_tmp_path)

        assert combined == [first_result, second_result]

    def test_filter_sheet_results_by_target_modes_keeps_requested_combination_only(
        self,
    ) -> None:
        requested_result = _build_sheet_result(
            model_type="cnn",
            recording_category="poem",
            norm_mode="norm_off",
            add_demo_data=False,
        )
        skipped_for_demo = _build_sheet_result(
            model_type="cnn",
            recording_category="a",
            norm_mode="norm_off",
            add_demo_data=True,
        )
        skipped_for_norm = _build_sheet_result(
            model_type="vit",
            recording_category="i",
            norm_mode="norm_on",
            add_demo_data=False,
        )
        skipped_for_both = _build_sheet_result(
            model_type="lstm",
            recording_category="o",
            norm_mode="norm_on",
            add_demo_data=True,
        )

        filtered_results, skipped_norm_count, skipped_demo_count = (
            filter_sheet_results_by_target_modes(
                sheet_results=[
                    requested_result,
                    skipped_for_demo,
                    skipped_for_norm,
                    skipped_for_both,
                ],
                norm_mode="norm_off",
                add_demo_data=False,
            )
        )

        assert filtered_results == [requested_result]
        assert skipped_norm_count == 2
        assert skipped_demo_count == 1

    def test_filter_sheet_results_by_target_modes_rejects_missing_demo_metadata(self) -> None:
        sheet_result = replace(
            _build_sheet_result(add_demo_data=False),
            add_demographic_data_var=None,
        )

        with pytest.raises(ValueError, match="add_demographic_data_var"):
            filter_sheet_results_by_target_modes(
                sheet_results=[sheet_result],
                norm_mode="norm_off",
                add_demo_data=False,
            )

    def test_filter_sheet_results_by_target_modes_raises_when_no_rows_match(self) -> None:
        sheet_results = [
            _build_sheet_result(norm_mode="norm_on", add_demo_data=False),
            _build_sheet_result(recording_category="a", norm_mode="norm_off", add_demo_data=True),
        ]

        with pytest.raises(ValueError, match="No loaded SheetResult entries matched"):
            filter_sheet_results_by_target_modes(
                sheet_results=sheet_results,
                norm_mode="norm_off",
                add_demo_data=False,
            )

    def test_build_excel_filename_uses_loaded_norm_and_demo_tokens(self) -> None:
        filename = build_excel_filename(
            model_variant="final",
            norm_mode="norm_on",
            add_demo_data=True,
            export_timestamp=EXPORT_TIMESTAMP,
        )

        assert (
            filename
            == "FINAL_MODEL_norm_on_w_demo_data_pickled_sheet_results_excel_summary_"
            "2026-03-19_14-05-33.xlsx"
        )

    def test_apply_duplicate_model_labels_renames_later_duplicate_identities(self) -> None:
        original = _build_sheet_result(model_type="cnn", add_demo_data=False)
        duplicate = _build_sheet_result(model_type="cnn", add_demo_data=False)

        prepared_results = apply_duplicate_model_labels([original, duplicate])

        assert prepared_results[0].model == "cnn"
        assert prepared_results[1].model == "Duplicate 2: cnn"
        assert original.model == "cnn"
        assert duplicate.model == "cnn"

    def test_save_excel_summaries_from_sheet_results_writes_combined_workbooks(
        self,
        local_tmp_path: Path,
    ) -> None:
        sheet_results = [
            _build_sheet_result(model_type="cnn", recording_category="poem", add_demo_data=False),
            _build_sheet_result(model_type="cnn", recording_category="poem", add_demo_data=False),
        ]

        final_path, best_path = save_excel_summaries_from_sheet_results(
            sheet_results=sheet_results,
            output_dir=local_tmp_path / "excel_output",
            norm_mode="norm_off",
            add_demo_data=False,
            export_timestamp=EXPORT_TIMESTAMP,
        )

        assert final_path == (
            local_tmp_path
            / "excel_output"
            / "FINAL_MODEL_norm_off_no_demo_data_pickled_sheet_results_excel_summary_2026-03-19_14-05-33.xlsx"
        )
        assert best_path == (
            local_tmp_path
            / "excel_output"
            / "BEST_MODEL_norm_off_no_demo_data_pickled_sheet_results_excel_summary_2026-03-19_14-05-33.xlsx"
        )
        assert final_path.exists()
        assert best_path.exists()

        workbook = load_workbook(final_path)
        sheet = workbook[workbook.sheetnames[0]]
        assert sheet["A1"].value == "Model"
        model_labels = {sheet["A2"].value, sheet["A3"].value}
        assert model_labels == {"cnn", "Duplicate 2: cnn"}

    def test_main_reads_load_sheet_results_dir_and_writes_filtered_target_output(
        self,
        local_tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        load_dir = local_tmp_path / "load_dir"
        output_dir = local_tmp_path / "output_dir"
        _write_pickle(
            load_dir / f"study_a{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}",
            [
                _build_sheet_result(model_type="cnn", norm_mode="norm_off", add_demo_data=False),
                _build_sheet_result(model_type="cnn", recording_category="a", norm_mode="norm_off", add_demo_data=True),
            ],
        )
        _write_pickle(
            load_dir / f"study_b{PATHS.SHEET_RESULTS_PICKLE_SUFFIX}",
            [
                _build_sheet_result(model_type="vit", recording_category="i", norm_mode="norm_on", add_demo_data=False),
                _build_sheet_result(model_type="lstm", recording_category="o", norm_mode="norm_on", add_demo_data=True),
            ],
        )

        monkeypatch.setattr(pickle_to_excel.PATHS, "LOAD_SHEET_RESULTS_DIR", load_dir)
        monkeypatch.setattr(
            pickle_to_excel,
            "get_year_month_day_hour_minute_second",
            lambda: EXPORT_TIMESTAMP,
        )

        cmd_args = SimpleNamespace(
            output_dir=str(output_dir),
            norm_mode="norm_off",
            add_demo_data=False,
        )

        exit_code = pickle_to_excel.main(cmd_args)

        assert exit_code == 0
        final_output = (
            output_dir
            / "FINAL_MODEL_norm_off_no_demo_data_pickled_sheet_results_excel_summary_2026-03-19_14-05-33.xlsx"
        )
        best_output = (
            output_dir
            / "BEST_MODEL_norm_off_no_demo_data_pickled_sheet_results_excel_summary_2026-03-19_14-05-33.xlsx"
        )
        assert final_output.exists()
        assert best_output.exists()

        workbook = load_workbook(final_output)
        sheet = workbook[workbook.sheetnames[0]]
        assert sheet["A2"].value == "cnn"
        assert sheet["A3"].value is None

    def test_cli_modes_include_pickle_to_excel(self) -> None:
        assert "pickle_to_excel" in CLI_CHOICES.MODES

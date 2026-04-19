from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Iterable, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from src.schemas.dataclasses import SheetResult



# -----------------------------
# helpers: formatting + color
# -----------------------------



RGB = Tuple[int, int, int]

# softer, less aggressive low-end
RED:    RGB = (0xF3, 0xB4, 0xB4)   # muted coral / dusty rose
ORANGE: RGB = (0xF2, 0xB1, 0x7A)   # F2B17A
YELLOW: RGB = (0xE9, 0xF2, 0x9B)   # lime-tinted yellow
GREEN:  RGB = (0x8F, 0xE6, 0xB3)   # slightly calmer green

def _clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))

def _hex_rgb(r: int, g: int, b: int) -> str:
    return f"{r:02X}{g:02X}{b:02X}"

def _lerp_color(c0: RGB, c1: RGB, t: float) -> RGB:
    t = _clamp(t, 0.0, 1.0)
    return tuple(int(round(a + (b - a) * t)) for a, b in zip(c0, c1))  # type: ignore[return-value]

def _fill_for_mean(mean: Optional[float]) -> Optional[PatternFill]:
    if mean is None:
        return None

    m = _clamp(float(mean), 0.0, 1.0)

    if m < 0.45:
        # red is rare
        t = m / 0.45
        rgb = _lerp_color(RED, ORANGE, t)

    elif m < 0.62:
        # orange/yellow zone is shorter
        t = (m - 0.45) / (0.62 - 0.45)
        rgb = _lerp_color(ORANGE, YELLOW, t)

    else:
        # green starts earlier and dominates more
        t = (m - 0.62) / (1.00 - 0.62)
        rgb = _lerp_color(YELLOW, GREEN, t)




    return PatternFill(fill_type="solid", fgColor=_hex_rgb(*rgb))


def _format_mean_pm_std(mean: Optional[float], std: Optional[float], decimals: int = 3) -> Optional[str]:
    if mean is None:
        return None
    if std is None:
        return f"{mean:.{decimals}f}"
    return f"{mean:.{decimals}f} (± {std:.{decimals}f})"


# ==================================
# Excel writer: group + pivot
# ==================================

def create_and_save_excel_file(
    results: Iterable["SheetResult"],
    output_dir: Union[str, Path],
    filename: str = "run_summary.xlsx",
    *,
    model_variant: str = "final",  # "final" or "best"
    best_configs_dict_path: Union[str, Path, None] = None,
    add_demo_data_column: bool = False,
) -> Path:
    """
    Creates an Excel workbook with:
      - 1 sheet per (dataset, lung_conditions)
      - rows = models
      - columns = rec_category blocks (val/test x MBA/AUROC)
      - values = string "mean (± std)" BUT fill color derived from mean only

    model_variant:
      - "final" -> uses SheetResult fields like final_model_val_mba_mean/std, ...
      - "best"  -> uses SheetResult fields like best_model_val_mba_mean/std, ...

    Optionally includes add_demo_data as its own column per recording-category block.
    A vertical separator column is inserted between recording-category blocks.
    If best_configs_dict_path is provided, a footer row is appended with that path.
    """
    if model_variant not in ("final", "best"):
        raise ValueError('model_variant must be "final" or "best"')

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename

    results = list(results)
    if not results:
        raise ValueError("create_and_save_excel_file(): got zero results")

    # --- strong vertical separator between recording-category blocks ---
    combo_separator_side = Side(style="thick", color="374151")

    # --- group results by sheet key ---
    def sheet_key(r: "SheetResult"):
        return (r.dataset, r.lung_conditions)

    groups = {}
    for r in results:
        k = sheet_key(r)
        if k not in groups:
            groups[k] = []
        groups[k].append(r)

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    tab_colors = ["4F81BD", "9BBB59", "C0504D", "8064A2", "4BACC6", "F79646"]

    for idx, (key, rows) in enumerate(sorted(groups.items(), key=lambda kv: kv[0])):
        dataset, lung_conditions = key

        base_name = "%s %s" % (dataset.upper(), lung_conditions)
        safe_name = (
            base_name.replace("/", "_")
            .replace("\\", "_")
            .replace(":", "_")
            .replace("*", "_")
            .replace("?", "_")
            .replace("[", "_")
            .replace("]", "_")
        )
        sheet_name = safe_name[:31]

        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]

        models = sorted(set(r.model for r in rows))
        rec_categories = sorted(set(r.recording_category for r in rows))

        # (model, rc) -> SheetResult
        lookup = {}
        for r in rows:
            lookup[(r.model, r.recording_category)] = r

        # --- build header ---
        header = ["Model"]
        col_specs = []  # list entries: (rec_category, base_metric)
        combo_boundary_columns = []

        block = [
            ("val_mba", "val MBA"),
            ("test_mba", "test MBA"),
            ("val_auroc", "val AUROC"),
            ("test_auroc", "test AUROC"),
        ]
        if add_demo_data_column:
            block.append(("add_demo_data", "add_demo_data"))
        for rc_idx, rc in enumerate(rec_categories):
            for base_metric, label in block:
                header.append("%s %s" % (rc.upper(), label))
                col_specs.append((rc, base_metric))
            if rc_idx < len(rec_categories) - 1:
                combo_boundary_columns.append(len(header))

        ws.append(header)

        # --- style header ---
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in combo_boundary_columns:
                cell.border = Border(right=combo_separator_side)

        # --- write data rows ---
        for model in models:
            row_vals = [model]

            # aligned to report columns
            per_cell_payload = []

            for rc, base_metric in col_specs:
                sr = lookup.get((model, rc))

                if base_metric == "add_demo_data":
                    add_demo_text = ""
                    if sr:
                        add_demo_value, _add_demo_status = sr.get_add_demographic_data_info()
                        add_demo_text = "True" if add_demo_value else "False"
                    row_vals.append(add_demo_text)
                    per_cell_payload.append(("add_demo_data", None))
                    continue

                # Pull either final_model_* or best_model_* fields
                mean_attr = "%s_model_%s_mean" % (model_variant, base_metric)
                std_attr = "%s_model_%s_std" % (model_variant, base_metric)

                mean_v = getattr(sr, mean_attr, None) if sr else None
                std_v = getattr(sr, std_attr, None) if sr else None

                row_vals.append(_format_mean_pm_std(mean_v, std_v, decimals=3))
                per_cell_payload.append(("metric", mean_v))

            ws.append(row_vals)
            written_row = ws.max_row

            # apply fills + alignment (report columns start at column 2)
            col_idx = 2
            for payload_kind, payload_value in per_cell_payload:
                cell = ws.cell(row=written_row, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if payload_kind == "metric":
                    fill = _fill_for_mean(payload_value)
                    if fill is not None:
                        cell.fill = fill
                if col_idx in combo_boundary_columns:
                    cell.border = Border(right=combo_separator_side)

                col_idx += 1

            # model column styling
            model_cell = ws.cell(row=written_row, column=1)
            model_cell.alignment = Alignment(horizontal="left", vertical="center")
            model_cell.font = Font(bold=True)

        if best_configs_dict_path is not None:
            footer_row = ws.max_row + 1
            footer_text = f"best_configs_dict_path: {Path(best_configs_dict_path)}"
            ws.cell(row=footer_row, column=1, value=footer_text)
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=ws.max_column)

            footer_font = Font(color="374151", italic=True, size=10)
            footer_fill = PatternFill("solid", fgColor="F9FAFB")
            footer_top_border = Border(top=Side(style="thin", color="D1D5DB"))

            for col in range(1, ws.max_column + 1):
                footer_cell = ws.cell(row=footer_row, column=col)
                footer_cell.font = footer_font
                footer_cell.fill = footer_fill
                if col in combo_boundary_columns:
                    footer_cell.border = Border(top=footer_top_border.top, right=combo_separator_side)
                else:
                    footer_cell.border = footer_top_border
                footer_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "B2"

        # autosize-ish columns
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 32)

        ws.cell(row=1, column=1).value = "Model"

    wb.save(out_path)
    return out_path



